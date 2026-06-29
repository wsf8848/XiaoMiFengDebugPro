#!/usr/bin/env python3
"""
小蜜蜂调试助手Pro - 授权码生成器（一机一码版）

功能：
1. 生成 200 个通用授权码（AES 加密后嵌入 APK）
2. 根据设备指纹生成专属授权码（一机一码，仅限特定设备使用）

一机一码使用流程：
  客户 → 打开 APP → 复制设备 ID 发给开发者
  开发者 → python gen_keys.py --device <设备ID> → 生成授权码
  客户 → 输入授权码 → 激活成功

HMAC 密钥必须与 LicenseChecker.kt 中的 buildHmacKey() 一致！
"""

import os, json, base64, secrets, hashlib, hmac, sys, argparse
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(OUT_DIR, "dist", "授权码列表_v1.5.0.xlsx")
CODES_KT_PATH = os.path.join(OUT_DIR, "app", "src", "main", "java", "com", "xmf", "debugpro", "Codes.kt")

# ═══ 一机一码 HMAC 密钥（必须与 LicenseChecker.kt 中的 buildHmacKey() 一致） ═══
_HMAC_PARTS = ["d4f7", "e2a9", "b1c8", "f30e"]
_HMAC_XOR = 0x5B

def build_hmac_key():
    raw = "".join(_HMAC_PARTS)
    key_bytes = bytes.fromhex(raw)
    return bytes(b ^ _HMAC_XOR for b in key_bytes)

def generate_device_code(device_id: str) -> str:
    """根据设备指纹生成一机一码授权码"""
    key = build_hmac_key()
    h = hmac.new(key, device_id.encode(), hashlib.sha256).hexdigest().upper()[:12]
    return f"{h[:4]}-{h[4:8]}-{h[8:12]}"

# ═══ 预生成码 AES 加密（保持不变） ═══
_KEY_PARTS = ["5e8f9a2b", "c7d3e1f4", "a6b9c0d2", "e3f7f818"]
_XOR_MASK = 0xA3

def build_key():
    raw = "".join(_KEY_PARTS)
    key_bytes = bytes.fromhex(raw)[:16]
    return bytes(b ^ _XOR_MASK for b in key_bytes)

def encrypt_codes(plain_list):
    key = build_key()
    iv = secrets.token_bytes(16)
    cipher = AES.new(key, AES.MODE_CBC, iv)
    plaintext = "\n".join(plain_list).encode("utf-8")
    ct = cipher.encrypt(pad(plaintext, AES.block_size))
    return base64.b64encode(iv + ct).decode("ascii")

# ═══ 主流程 ═══
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="授权码生成器")
    parser.add_argument("--device", help="设备指纹（16位大写HEX），生成该设备的专属一机一码")
    args = parser.parse_args()

    # ── 一机一码模式 ──
    if args.device:
        code = generate_device_code(args.device)
        print("═" * 50)
        print("  一机一码 授权码")
        print("═" * 50)
        print()
        print(f"  设备指纹：{args.device}")
        print(f"  授权码：  {code}")
        print()
        print("═" * 50)
        print("将该授权码发给客户，客户在APP中输入即可激活。")
        sys.exit(0)

    # ── 批量生成模式（200个通用码） ──
    def generate_code(index):
        rand = secrets.token_hex(6).upper()
        return f"{rand[:4]}-{rand[4:8]}-{rand[8:]}"

    codes = []
    code_set = set()
    while len(codes) < 200:
        c = generate_code(len(codes) + 1)
        if c not in code_set:
            code_set.add(c)
            codes.append(c)
    print(f"✅ 已生成 {len(codes)} 个不重复授权码")

    # Excel
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "授权码列表"
    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    for col, title in enumerate(["序号", "授权码", "状态", "备注"], 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    avail_font = Font(name="微软雅黑", size=11, color="006600")
    avail_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    for i, code in enumerate(codes):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=code).alignment = Alignment(horizontal="center")
        status_cell = ws.cell(row=row, column=3, value="未售")
        status_cell.font = avail_font; status_cell.fill = avail_fill
        status_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value="")
    wb.save(EXCEL_PATH)
    print(f"📊 Excel: {EXCEL_PATH}")

    # Codes.kt
    encrypted = encrypt_codes(codes)
    segment_size = len(encrypted) // 5
    segments = [encrypted[i * segment_size: (i + 1) * segment_size if i < 4 else len(encrypted)] for i in range(5)]

    kt_code = f"""/**
 * 授权码加密数据 - 自动生成，请勿手动修改
 * 生成时间: 2026-06-29
 * 数量: {len(codes)} 个
 */
package com.xmf.debugpro

object Codes {{
    private val _s0 = "{segments[0]}"
    private val _s1 = "{segments[1]}"
    private val _s2 = "{segments[2]}"
    private val _s3 = "{segments[3]}"
    private val _s4 = "{segments[4]}"

    val encryptedData: String get() = _s0 + _s1 + _s2 + _s3 + _s4
}}
"""
    with open(CODES_KT_PATH, "w", encoding="utf-8") as f:
        f.write(kt_code)
    print(f"🔐 Codes.kt: {CODES_KT_PATH}")
    print(f"\n生成完毕！前5个授权码：")
    for i in range(5):
        print(f"  {i+1}. {codes[i]}")
